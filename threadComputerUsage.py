import time

from openpyxl import *
from openpyxl.styles import *
from concurrent.futures import ThreadPoolExecutor
from computerUsage import ComputerUsage
from sqlExecute import SqlExecute
from noSqlExecute import NoSqlExecute
import datetime
import os
from pyspark.sql import SparkSession

# Start master: spark-class org.apache.spark.deploy.master.Master
# Start worker: spark-class org.apache.spark.deploy.worker.Worker -c 4 -m 4G spark://192.168.0.166:7077


def init(font, alignment):
    # Start Spark app
    os.environ['PYSPARK_SUBMIT_ARGS'] = '--packages com.datastax.spark:spark-cassandra-connector_2.12:3.1.0 ' \
                                     '--conf spark.cassandra.connection.host=127.0.0.1 pyspark-shell'

    spark = SparkSession.builder.appName("SQL Example").master("local[*]") \
        .config("spark.sql.catalog.history", "com.datastax.spark.connector.datasource.CassandraCatalog") \
        .config("spark.cassandra.connection.host", "127.0.0.1") \
        .config("spark.sql.extensions", "com.datastax.spark.connector.CassandraSparkExtensions") \
        .getOrCreate()

    spark.read.format("org.apache.spark.sql.cassandra").option("table", "customer").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("customer")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "order_sales").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("order_sales")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "payment").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("payment")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "product").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("product")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "review").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("review")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "sale").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("sale")
    spark.read.format("org.apache.spark.sql.cassandra").option("table", "seller").option("keyspace", "dw4cassandra") \
        .load().createOrReplaceTempView("seller")

    wb_destination = Workbook()
    ws_destination = wb_destination.active
    ws_destination.row_dimensions[1].height = ws_destination.row_dimensions[2].height = 20
    ws_destination.column_dimensions['A'].width = ws_destination.column_dimensions['B'].width = \
        ws_destination.column_dimensions['C'].width = ws_destination.column_dimensions['D'].width = \
        ws_destination.column_dimensions['E'].width = ws_destination.column_dimensions['F'].width = \
        ws_destination.column_dimensions['G'].width = ws_destination.column_dimensions['H'].width = \
        ws_destination.column_dimensions['I'].width = ws_destination.column_dimensions['J'].width = 15
    ws_destination.column_dimensions['K'].width = 260
    ws_destination.title = "Mysql and Cassandra Queries"
    ws_destination.merge_cells('A1:E1')
    ws_destination.merge_cells('F1:J1')
    ws_destination['A1'].font = ws_destination['A2'].font = ws_destination['B2'].font = \
        ws_destination['C2'].font = ws_destination['D2'].font = ws_destination['E2'].font = \
        ws_destination['F1'].font = ws_destination['F2'].font = ws_destination['G2'].font = \
        ws_destination['H2'].font = ws_destination['I2'].font = ws_destination['J2'].font = \
        ws_destination['K1'].font = ws_destination['K2'].font = font
    ws_destination['A1'].alignment = ws_destination['A2'].alignment = ws_destination['B2'].alignment = \
        ws_destination['C2'].alignment = ws_destination['D2'].alignment = ws_destination['E2'].alignment = \
        ws_destination['F1'].alignment = ws_destination['F2'].alignment = ws_destination['G2'].alignment = \
        ws_destination['H2'].alignment = ws_destination['I2'].alignment = ws_destination['J2'].alignment = \
        ws_destination['K1'].alignment = ws_destination['K2'].alignment = alignment
    ws_destination['A1'] = "Mysql"
    ws_destination['F1'] = "Cassandra"
    ws_destination['K1'] = "Queries Used"
    ws_destination['A2'] = ws_destination['F2'] = "Time used"
    ws_destination['B2'] = ws_destination['G2'] = "Bytes used"
    ws_destination['C2'] = ws_destination['H2'] = "CPU used"
    ws_destination['D2'] = ws_destination['I2'] = "RAM used"
    ws_destination['E2'] = ws_destination['J2'] = "SWAP used"
    ws_destination['K2'] = "Query"

    return wb_destination, ws_destination,  spark


with ThreadPoolExecutor() as executor:
    font = Font(name="Times", size=14)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb_results, ws_results, spark = init(font=font, alignment=alignment)

    wb_queries = load_workbook(filename="geradorConsulta.xlsx", data_only=True)
    ws_queries = wb_queries.active

    usage = ComputerUsage()
    sqlExec = SqlExecute()
    noSqlExec = NoSqlExecute(spark=spark)

    line = 3
    for j in range(4, ws_queries.max_column + 1):
        for i in range(4, ws_queries.max_row + 1):
            if ws_queries.cell(row=i, column=j).value is not None:
                print("Now = %s, i = %d, j = %d" % (datetime.datetime.now(), i, j))
                noSqlExec.query = sqlExec.sql = ws_queries.cell(row=i, column=j).value

                usage.start()
                main_thread = executor.submit(usage.run)
                try:
                    new_thread = executor.submit(sqlExec.run)
                finally:
                    time_spent, bytes_usage = new_thread.result()
                    usage.keep_measuring = False
                    cpu_usage, ram_usage, swap_usage = main_thread.result()
                    ws_results['A' + str(line)].font = ws_results['B' + str(line)].font = \
                        ws_results['C' + str(line)].font = ws_results['D' + str(line)].font = \
                        ws_results['E' + str(line)].font = font
                    ws_results['A' + str(line)].alignment = ws_results['B' + str(line)].alignment = \
                        ws_results['C' + str(line)].alignment = ws_results['D' + str(line)].alignment = \
                        ws_results['E' + str(line)].alignment = alignment
                    ws_results['A' + str(line)] = time_spent
                    ws_results['B' + str(line)] = bytes_usage
                    ws_results['C' + str(line)] = str(cpu_usage)[1:-1]
                    ws_results['D' + str(line)] = str(ram_usage)[1:-1]
                    ws_results['E' + str(line)] = str(swap_usage)[1:-1]

                time.sleep(5)
                usage.start()
                main_thread = executor.submit(usage.run)
                try:
                    new_thread = executor.submit(noSqlExec.run)
                finally:
                    time_spent, bytes_usage = new_thread.result()
                    usage.keep_measuring = False
                    cpu_usage, ram_usage, swap_usage = main_thread.result()
                    ws_results['F' + str(line)].font = ws_results['G' + str(line)].font = \
                        ws_results['H' + str(line)].font = ws_results['I' + str(line)].font = \
                        ws_results['J' + str(line)].font = font
                    ws_results['F' + str(line)].alignment = ws_results['G' + str(line)].alignment = \
                        ws_results['H' + str(line)].alignment = ws_results['I' + str(line)].alignment = \
                        ws_results['J' + str(line)].alignment = alignment
                    ws_results['F' + str(line)] = time_spent
                    ws_results['G' + str(line)] = bytes_usage
                    ws_results['H' + str(line)] = str(cpu_usage)[1:-1]
                    ws_results['I' + str(line)] = str(ram_usage)[1:-1]
                    ws_results['J' + str(line)] = str(swap_usage)[1:-1]

                ws_results['K' + str(line)].font = font
                ws_results['K' + str(line)].alignment = alignment
                ws_results['K' + str(line)] = ws_queries.cell(row=i, column=j).value
                ws_results.row_dimensions[line].height = 20
                line += 1
                time.sleep(5)

            else:
                break

    wb_results.save(filename="queryResults.xlsx")
