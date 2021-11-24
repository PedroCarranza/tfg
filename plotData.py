import matplotlib.pyplot as plt
from openpyxl import *
from matplotlib import *
from datetime import *


def print_options():
    print("1 - Time used to make the query \n"
          "2 - Bytes used to make the query \n"
          "3 - CPU used to make the query \n"
          "4 - RAM used to make the query \n"
          "5 - SWAP used to make the query \n"
          "6 - Exit\n")


workbook = load_workbook(filename="queryResults.xlsx", data_only=True)
worksheet = workbook.active

mysql_values = []
cassandra_values = []
for i in range(3, worksheet.max_row + 1):
    mysql_values.append({'query': str(worksheet.cell(row=i, column=11).value),
                        'time_used': (datetime.combine(date.min,
                                                       worksheet.cell(row=i, column=1).value) - datetime.min)
                         .total_seconds(),
                         'bytes_used': int(worksheet.cell(row=i, column=2).value) / 1024 / 1024,
                         'cpu_used': sum(map(float, (str(worksheet.cell(row=i, column=3).value).split(",")))) /
                         len(map(float, (str(worksheet.cell(row=i, column=3).value).split(",")))),
                         'ram_used': sum(map(float, (str(worksheet.cell(row=i, column=4).value).split(",")))) /
                         len(map(float, (str(worksheet.cell(row=i, column=4).value).split(",")))) / 1024 / 1024,
                         # Get MB value
                         'swap_used': sum(map(float, (str(worksheet.cell(row=i, column=5).value).split(",")))) /
                         len(map(float, (str(worksheet.cell(row=i, column=5).value).split(",")))) / (10**6)
                         # Get MB value
                         })
    cassandra_values.append({'query': str(worksheet.cell(row=i, column=11).value),
                             'time_used': (datetime.combine(date.min,
                                                            worksheet.cell(row=i, column=6).value) - datetime.min)
                             .total_seconds(),
                             'bytes_used': int(worksheet.cell(row=i, column=7).value) / 1024 / 1024,
                             'cpu_used': sum(map(float, (str(worksheet.cell(row=i, column=8).value).split(",")))) /
                             len(map(float, (str(worksheet.cell(row=i, column=8).value).split(",")))),
                             'ram_used': sum(map(float, (str(worksheet.cell(row=i, column=9).value).split(",")))) /
                             len(map(float, (str(worksheet.cell(row=i, column=9).value).split(",")))) / 1024 / 1024,
                             # Get MB value
                             'swap_used': sum(map(float, (str(worksheet.cell(row=i, column=10).value).split(",")))) /
                             len(map(float, (str(worksheet.cell(row=i, column=10).value).split(",")))) / (10**6)
                             # Get MB value
                             })
cassandra_values.sort(key=lambda d: d['time_used'])
queries_file = open('queries.txt', 'w')
query_order_values = [cassandra_values[value]['query'] for value in range(0, len(cassandra_values))]
for i in range(0, len(query_order_values)):
    queries_file.write('\t\t' + str(i+1) + ' & ' + query_order_values[i].replace('_', '$\\_$') + ' \\\\\n')
queries_file.close()

mysql_values_aux = mysql_values
mysql_values = []
for i in range(0, len(cassandra_values)):
    for j in range(0, len(mysql_values_aux)):
        if mysql_values_aux[j]['query'] == cassandra_values[i]['query']:
            mysql_values.append(mysql_values_aux[j])
            break

print_options()
option = int(input("Select an option: "))

while 1 <= option <= 5 or option != 6:
    n = [i for i in range(1, worksheet.max_row - 1)]

    if option == 1:
        field_to_plot = 'time_used'
        plt.ylabel("Tempo gasto por consulta (s)")
    elif option == 2:
        field_to_plot = 'bytes_used'
        plt.ylabel("Bytes em disco usados por consulta (MB)")
    elif option == 3:
        field_to_plot = 'cpu_used'
        plt.ylabel("Uso de cpu por consulta (%)")
    elif option == 4:
        field_to_plot = 'ram_used'
        plt.ylabel("Uso de ram por consulta (MB)")
    elif option == 5:
        field_to_plot = 'swap_used'
        plt.ylabel("Swaps feitos por consulta (10^6)")

    mysql_sub_values = [mysql_values[value][field_to_plot] for value in range(0, len(mysql_values))]
    cassandra_sub_values = [cassandra_values[value][field_to_plot] for value in range(0, len(cassandra_values))]
    print(mysql_sub_values)
    print(cassandra_sub_values)
    plt.plot(n, mysql_sub_values, 'b', label="MySQL")
    plt.plot(n, cassandra_sub_values, 'r', label="Cassandra")
    plt.axis([0, worksheet.max_row-1, -5, max(max(mysql_sub_values), max(cassandra_sub_values))])
    plt.xlabel("Numero da consulta conforme anexo")
    plt.legend()
    plt.show()
    option = int(input("Select an option: "))
